""" 학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
1,10,8,5,14,26,12
2,7,3,7,15,24,18
3,9,5,8,8,12,4
4,7,8,7,17,21,18
5,7,8,7,16,25,15
6,3,5,8,8,17,0
7,4,9,10,16,27,18
8,6,6,6,15,19,17
9,10,10,9,19,30,19
10,9,8,8,20,25,20 """

from openpyxl import Workbook
wb = Workbook()
ws = wb.active

#데이터를 넣기
ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))  # tuple형태로 넣어줌

scores = [
    (1, 10, 8, 5, 14, 26, 12),
    (2, 7, 3, 7, 15, 24, 18),
    (3, 9, 5, 8, 8, 12, 4),
    (4, 7, 8, 7, 17, 21, 18),
    (5, 7, 8, 7, 16, 25, 15),
    (6, 3, 5, 8, 8, 17, 0),
    (7, 4, 9, 10, 16, 27, 18),
    (8, 6, 6, 6, 15, 19, 17),
    (9, 10, 10, 9, 19, 30, 19),
    (10, 9, 8, 8, 20, 25, 20)
]

for s in scores:
    ws.append(s)

#퀴즈 2를 10점으로 변경
for idx, cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    cell.value = 10

ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:])-score[3]+10 #엑셀 수식을 활용한 것은 아직 evaluate되기 전이라서 사용하는데에 문제가 있어서 이렇게 따로 변수를 만들어준 것이다.
    ws.cell(idx, 8).value = "=SUM(B{}:G{})".format(idx, idx)
    if score[1] < 5:
        ws.cell(idx, 9).value = "F"
    elif sum_val >= 90:
        ws.cell(idx, 9).value = "A"
    elif sum_val >= 80:
        ws.cell(idx, 9).value = "B"
    elif sum_val >= 70:
        ws.cell(idx, 9).value = "C"
    else:
        ws.cell(idx, 9).value = "D"
    

wb.save(".\RPA\Excel\ExcelFiles\scores_1.xlsx")
