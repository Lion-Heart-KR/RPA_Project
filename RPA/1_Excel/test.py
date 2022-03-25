from openpyxl import load_workbook
wb = load_workbook(".\RPA\Excel\ExcelFiles\scores_1.xlsx")
ws = wb.active

print(ws["1"][1:])