from openpyxl import Workbook, load_workbook

# wb = Workbook()
# ws = wb.active

# # initailize data
# for i in range(1, 12):
#     rows = input("성적을 입력해주세요").split(',')
#     ws.append(rows)

wb = load_workbook(".\RPA\Excel\ExcelFiles\scores.xlsx", data_only=True)
ws = wb.active

# 퀴즈 2 점수를 10점으로 수정
# for col in range(1, ws.max_column+1):
#     if "퀴즈2" in ws.cell(1, col).value:
#         for row in range(1, ws.max_row):
#             ws.cell(row+1, col).value=10

#         break

# # #H 열에 총점(SUM함수 이용해서 넣기)
# col = ws.max_column+1
# ws.cell(1, col, value="총점")
# for row in range(1,ws.max_row):
#     ws.cell(row+1, col).value = "=SUM(B"+str(row+1)+":G"+str(row+1)+")"


# #I열에 성적 정보 추가
col = ws.max_column+1
ws.cell(1, col, value="성적")
for row in range(1,ws.max_row):
    if ws.cell(row+1, 2).value < 5:
        ws.cell(row+1, col).value = "F"
    elif ws.cell(row+1, col-1).value >= 90:
        ws.cell(row+1, col).value = "A"
    elif ws.cell(row+1, col-1).value >= 80:
        ws.cell(row+1, col).value = "B"
    elif ws.cell(row+1, col-1).value >= 70:
        ws.cell(row+1, col).value = "C"
    else:
        ws.cell(row+1, col).value = "D"

wb.save(".\RPA\Excel\ExcelFiles\scores.xlsx")